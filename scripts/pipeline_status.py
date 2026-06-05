#!/usr/bin/env python3
"""Report pipeline state and what's pending at each stage.

Stages:
  1. Enrichment   — input CSV IDs vs IDs present in enriched batches / final XLSX
  2. ID→slug map  — does state/id_slug_map.json reflect the enriched file?
  3. Prompts      — does each ID have a rendered prompt?
  4. Analyses     — does each prompt have a written analysis?
  5. Compile      — are all analyses included in the compiled .md?

Run with `python scripts/pipeline_status.py` to see the punch list.
"""
import csv
import json
import re
from pathlib import Path

import openpyxl

from _paths import (
    INPUT_CSV, ENRICHED_BATCHES_DIR, ENRICHED_FINAL,
    EXEC_PROMPTS_DIR, EXEC_ANALYSES_DIR, COMPILED_MD,
    ID_SLUG_MAP, COMPILED_STATE,
    DESC_PROMPTS_DIR, DESC_FICHAS_DIR, DESC_COMPILED_MD, DESC_COMPILED_STATE,
)


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (text or "").lower()).strip("_")


def input_ids() -> set[str]:
    if not INPUT_CSV.exists():
        return set()
    with INPUT_CSV.open(encoding="utf-8-sig") as f:
        return {row["Identifier"].strip() for row in csv.DictReader(f) if row.get("Identifier")}


def input_names() -> dict[str, str]:
    if not INPUT_CSV.exists():
        return {}
    with INPUT_CSV.open(encoding="utf-8-sig") as f:
        return {row["Identifier"].strip(): row["Name"].strip()
                for row in csv.DictReader(f)
                if row.get("Identifier") and row.get("Name")}


def ids_in_xlsx(path: Path) -> set[str]:
    if not path.exists():
        return set()
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return set()
    if "AI Features & Agents" not in wb.sheetnames:
        return set()
    ws = wb["AI Features & Agents"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None) or ()
    if "Identifier" not in header:
        return set()
    idx = header.index("Identifier")
    return {r[idx] for r in rows if r and r[idx]}


def ids_in_batches() -> set[str]:
    out: set[str] = set()
    if not ENRICHED_BATCHES_DIR.exists():
        return out
    for p in ENRICHED_BATCHES_DIR.glob("*.xlsx"):
        out |= ids_in_xlsx(p)
    return out


def main() -> int:
    in_ids = input_ids()
    names = input_names()
    print(f"Input CSV:                    {INPUT_CSV}")
    print(f"  Total IDs in input:         {len(in_ids)}")

    batch_ids = ids_in_batches()
    final_ids = ids_in_xlsx(ENRICHED_FINAL)
    print()
    print("Stage 1 — Enrichment")
    print(f"  Batches dir:                {ENRICHED_BATCHES_DIR}")
    print(f"  IDs covered by batches:     {len(batch_ids)}")
    print(f"  IDs in final XLSX:          {len(final_ids)}  ({ENRICHED_FINAL.name})")
    pending_enrich = in_ids - (batch_ids | final_ids)
    print(f"  Pending enrichment:         {len(pending_enrich)}")
    if pending_enrich:
        sample = ", ".join(sorted(pending_enrich)[:10])
        more = f" (+{len(pending_enrich)-10} more)" if len(pending_enrich) > 10 else ""
        print(f"    e.g. {sample}{more}")

    print()
    print("Stage 2 — ID→slug map")
    if ID_SLUG_MAP.exists():
        m = json.loads(ID_SLUG_MAP.read_text(encoding="utf-8"))
        print(f"  {ID_SLUG_MAP} exists with {len(m)} entries")
        missing_in_map = (batch_ids | final_ids) - set(m.keys())
        if missing_in_map:
            print(f"  Out of sync — missing in map: {len(missing_in_map)}")
    else:
        print(f"  {ID_SLUG_MAP} does NOT exist — run scripts/build_id_slug_map.py")

    print()
    print("Stage 3 — Prompts")
    existing_prompts = {p.name for p in EXEC_PROMPTS_DIR.glob("*.prompt.md")}
    pending_prompts = []
    for code in (batch_ids | final_ids):
        name = names.get(code)
        if not name:
            continue
        expected = f"{code.lower()}_{slugify(name)}_analysis.prompt.md"
        if expected not in existing_prompts:
            pending_prompts.append((code, expected))
    print(f"  Rendered prompts:           {len(existing_prompts)}")
    print(f"  Pending prompts:            {len(pending_prompts)}")
    if pending_prompts:
        for code, fn in pending_prompts[:5]:
            print(f"    {code} -> {fn}")

    print()
    print("Stage 4 — Analyses")
    existing_analyses = {p.name for p in EXEC_ANALYSES_DIR.glob("*.md")}
    pending_analyses = []
    for prompt_name in existing_prompts:
        analysis_name = prompt_name.replace(".prompt.md", ".md")
        if analysis_name not in existing_analyses:
            pending_analyses.append(analysis_name)
    print(f"  Written analyses:           {len(existing_analyses)}")
    print(f"  Pending analyses:           {len(pending_analyses)}")
    if pending_analyses:
        for a in pending_analyses[:5]:
            print(f"    {a}")

    print()
    print("Stage 5 — Compiled RAG file")
    if COMPILED_STATE.exists():
        compiled = set(json.loads(COMPILED_STATE.read_text(encoding="utf-8")).get("compiled_files", []))
    else:
        compiled = set()
    pending_compile = sorted(existing_analyses - compiled)
    print(f"  Compiled file:              {COMPILED_MD}  {'(exists)' if COMPILED_MD.exists() else '(missing)'}")
    print(f"  Already compiled:           {len(compiled)}")
    print(f"  Pending compile:            {len(pending_compile)}")
    if pending_compile:
        for a in pending_compile[:5]:
            print(f"    {a}")

    # --- Description pipeline (fichas explicativas) ---
    print()
    print("Stage 6 — Description prompts")
    existing_desc_prompts = {p.name for p in DESC_PROMPTS_DIR.glob("*.prompt.md")}
    pending_desc_prompts = []
    for code in (batch_ids | final_ids):
        name = names.get(code)
        if not name:
            continue
        expected = f"{code.lower()}_{slugify(name)}_description.prompt.md"
        if expected not in existing_desc_prompts:
            pending_desc_prompts.append((code, expected))
    print(f"  Rendered description prompts: {len(existing_desc_prompts)}")
    print(f"  Pending description prompts:  {len(pending_desc_prompts)}")
    if pending_desc_prompts:
        for code, fn in pending_desc_prompts[:5]:
            print(f"    {code} -> {fn}")

    print()
    print("Stage 7 — Fichas (descriptions)")
    existing_fichas = {p.name for p in DESC_FICHAS_DIR.glob("*.md")}
    pending_fichas = []
    for prompt_name in existing_desc_prompts:
        ficha_name = prompt_name.replace(".prompt.md", ".md")
        if ficha_name not in existing_fichas:
            pending_fichas.append(ficha_name)
    print(f"  Written fichas:               {len(existing_fichas)}")
    print(f"  Pending fichas:               {len(pending_fichas)}")
    if pending_fichas:
        for a in pending_fichas[:5]:
            print(f"    {a}")

    print()
    print("Stage 8 — Descriptions compiled RAG file")
    if DESC_COMPILED_STATE.exists():
        desc_compiled = set(json.loads(DESC_COMPILED_STATE.read_text(encoding="utf-8")).get("compiled_files", []))
    else:
        desc_compiled = set()
    pending_desc_compile = sorted(existing_fichas - desc_compiled)
    print(f"  Compiled file:                {DESC_COMPILED_MD}  {'(exists)' if DESC_COMPILED_MD.exists() else '(missing)'}")
    print(f"  Already compiled:             {len(desc_compiled)}")
    print(f"  Pending compile:              {len(pending_desc_compile)}")

    print()
    print("Summary:")
    print(f"  pending enrichment:        {len(pending_enrich)}")
    print(f"  pending prompts:           {len(pending_prompts)}")
    print(f"  pending analyses:          {len(pending_analyses)}")
    print(f"  pending compile:           {len(pending_compile)}")
    print(f"  pending description prompts:{len(pending_desc_prompts)}")
    print(f"  pending fichas:            {len(pending_fichas)}")
    print(f"  pending description compile:{len(pending_desc_compile)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
