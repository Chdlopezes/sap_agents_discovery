#!/usr/bin/env python3
"""Persist one enrichment batch as a 3-sheet XLSX.

The enrichment agent calls this with a JSON payload containing the data it
extracted from SAP Discovery Center Detail Pages, and this script writes the
canonical XLSX into `enriched_data/batches/`.

Payload shape (JSON):
{
  "start": 1,                       # first record number in the batch (1-based)
  "end":   10,                      # last record number in the batch
  "ai_features_agents": [           # one row per record in the batch
    {
      "Name": "...", "AI Type": "AI Feature", "Commercial Type": "Base",
      "Product": "...", "Description": "...", "Product Category": "...",
      "Package": "...", "Quick Filters": "...", "Availability": "...",
      "Identifier": "J597", "Detail Page": "https://...",
      "Overview": "...", "Beneficios": "...", "Business Value": "..."
    },
    ...
  ],
  "pricing_premium": [              # only Premium records (may be empty)
    {
      "Name": "...", "Product": "...", "Commercial Type": "Premium",
      "Package": "...", "Identifier": "...", "Detail Page": "...",
      "Pricing Details": "..."
    },
    ...
  ],
  "initial_setup": [                # one row per record in the batch
    {
      "Name": "...", "Product": "...", "Identifier": "...", "Detail Page": "...",
      "Prerequisitos": "...", "Procedimiento": "...", "Próximos Pasos": "...",
      "Dificultad estimada": "Baja|Media|Alta",
      "Tipo": "...",
      "Link": "..."
    },
    ...
  ]
}

Usage:
  python scripts/save_enriched_batch.py --json batch.json
  python scripts/save_enriched_batch.py --stdin   # read JSON from stdin

Output: enriched_data/batches/AI_Features_Data_lote_<start>_<end>_enriquecido.xlsx
"""
import argparse
import json
import sys
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from _paths import ENRICHED_BATCHES_DIR

SHEETS = (
    ("AI Features & Agents", "ai_features_agents", [
        "Name", "AI Type", "Commercial Type", "Product", "Description",
        "Product Category", "Package", "Quick Filters", "Availability",
        "Identifier", "Detail Page", "Overview", "Beneficios", "Business Value",
    ]),
    ("Pricing Premium", "pricing_premium", [
        "Name", "Product", "Commercial Type", "Package", "Identifier",
        "Detail Page", "Pricing Details",
    ]),
    ("Initial Setup", "initial_setup", [
        "Name", "Product", "Identifier", "Detail Page",
        "Prerequisitos", "Procedimiento", "Próximos Pasos",
        "Dificultad estimada", "Tipo", "Link",
    ]),
)


def _write_sheet(ws, columns: list[str], rows: list[dict[str, Any]]):
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=c, value=name)
        cell.font = bold
    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(columns, start=1):
            cell = ws.cell(row=r, column=c, value=row.get(name, ""))
            cell.alignment = wrap
    # Reasonable column widths
    widths = {
        "Name": 38, "AI Type": 14, "Commercial Type": 16, "Product": 36,
        "Description": 60, "Product Category": 28, "Package": 36,
        "Quick Filters": 18, "Availability": 18, "Identifier": 12,
        "Detail Page": 60, "Overview": 70, "Beneficios": 70,
        "Business Value": 60, "Pricing Details": 70,
        "Prerequisitos": 60, "Procedimiento": 70, "Próximos Pasos": 50,
        "Dificultad estimada": 18, "Tipo": 28, "Link": 60,
    }
    for c, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(name, 24)
    ws.freeze_panes = "A2"


def write_batch(payload: dict[str, Any]) -> Path:
    start = int(payload["start"])
    end = int(payload["end"])
    ENRICHED_BATCHES_DIR.mkdir(parents=True, exist_ok=True)
    out = ENRICHED_BATCHES_DIR / f"AI_Features_Data_lote_{start}_{end}_enriquecido.xlsx"
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_title, key, columns in SHEETS:
        rows = payload.get(key) or []
        ws = wb.create_sheet(title=sheet_title)
        _write_sheet(ws, columns, rows)
    wb.save(out)
    return out


def main() -> int:
    p = argparse.ArgumentParser()
    src = p.add_mutually_exclusive_group(required=True)
    src.add_argument("--json", type=Path, help="Path to JSON file with the batch payload")
    src.add_argument("--stdin", action="store_true", help="Read JSON payload from stdin")
    args = p.parse_args()

    if args.stdin:
        payload = json.load(sys.stdin)
    else:
        payload = json.loads(args.json.read_text(encoding="utf-8"))

    out = write_batch(payload)
    print(f"Wrote {out}")
    print(f"  AI Features & Agents rows: {len(payload.get('ai_features_agents') or [])}")
    print(f"  Pricing Premium rows:      {len(payload.get('pricing_premium') or [])}")
    print(f"  Initial Setup rows:        {len(payload.get('initial_setup') or [])}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
