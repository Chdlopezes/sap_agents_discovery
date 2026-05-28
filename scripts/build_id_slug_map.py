#!/usr/bin/env python3
"""Build state/id_slug_map.json from the enriched XLSX.

The map is { "<Identifier>": "<slug>" }, where slug is slugify(Name).
Used by downstream consumers (e.g. RAG indexing) that need the slug
without having to parse the XLSX.

Usage:
  python scripts/build_id_slug_map.py
"""
import json
import re

import openpyxl

from _paths import ENRICHED_FINAL, ID_SLUG_MAP, STATE_DIR


def slugify(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (text or "").lower()).strip("_")


def main() -> int:
    if not ENRICHED_FINAL.exists():
        print(f"ERROR: {ENRICHED_FINAL} does not exist. Run scripts/merge_enriched_batches.py first.")
        return 1
    wb = openpyxl.load_workbook(ENRICHED_FINAL, read_only=True, data_only=True)
    if "AI Features & Agents" not in wb.sheetnames:
        print("ERROR: enriched XLSX is missing the 'AI Features & Agents' sheet.")
        return 2
    ws = wb["AI Features & Agents"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    id_idx = header.index("Identifier")
    name_idx = header.index("Name")
    mapping: dict[str, str] = {}
    for r in rows:
        if not r:
            continue
        code = r[id_idx]
        name = r[name_idx]
        if not code or not name:
            continue
        mapping[str(code).strip()] = slugify(str(name))
    STATE_DIR.mkdir(parents=True, exist_ok=True)
    ID_SLUG_MAP.write_text(
        json.dumps(mapping, ensure_ascii=False, indent=2, sort_keys=True),
        encoding="utf-8",
    )
    print(f"Wrote {ID_SLUG_MAP} with {len(mapping)} entries.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
