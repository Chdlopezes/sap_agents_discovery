#!/usr/bin/env python3
"""Merge all per-batch enriched XLSX files into the final enriched XLSX.

- Reads every `enriched_data/batches/*.xlsx`.
- Concatenates rows across batches, sheet by sheet (AI Features & Agents,
  Pricing Premium, Initial Setup).
- Deduplicates by Identifier within each sheet; on conflict, prefers the most
  recent batch (last write wins).
- Writes `processed/AI_Features_Data_Enriched.xlsx` with frozen header row.

Idempotent: safe to run repeatedly.

Usage:
  python scripts/merge_enriched_batches.py
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from _paths import ENRICHED_BATCHES_DIR, ENRICHED_FINAL, ENRICHED_SHEETS


def _read_sheet(path: Path, sheet: str) -> tuple[list[str], list[dict]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet not in wb.sheetnames:
        return [], []
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        return [], []
    header = [h for h in header if h is not None]
    out_rows: list[dict] = []
    for r in rows:
        if not r or all(v in (None, "") for v in r):
            continue
        out_rows.append({header[i]: r[i] for i in range(len(header)) if i < len(r)})
    return header, out_rows


def _merge_sheet(sheet: str, batch_files: list[Path]) -> tuple[list[str], list[dict]]:
    columns_union: list[str] = []
    by_id: dict[str, dict] = {}
    no_id_rows: list[dict] = []
    for bf in batch_files:
        header, rows = _read_sheet(bf, sheet)
        for h in header:
            if h not in columns_union:
                columns_union.append(h)
        for row in rows:
            ident = (row.get("Identifier") or "").strip() if row.get("Identifier") else ""
            if ident:
                by_id[ident] = row  # last batch wins
            else:
                no_id_rows.append(row)
    merged = list(by_id.values()) + no_id_rows
    # Stable order: by Identifier if present
    merged.sort(key=lambda r: (r.get("Identifier") is None, str(r.get("Identifier") or "")))
    return columns_union, merged


def _write_sheet(ws, columns: list[str], rows: list[dict]):
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    for c, name in enumerate(columns, start=1):
        ws.cell(row=1, column=c, value=name).font = bold
    for r, row in enumerate(rows, start=2):
        for c, name in enumerate(columns, start=1):
            val = row.get(name)
            ws.cell(row=r, column=c, value=val if val is not None else "").alignment = wrap
    for c, name in enumerate(columns, start=1):
        # Heuristic width; long text columns get more space.
        long_cols = {"Overview", "Beneficios", "Business Value", "Pricing Details",
                     "Prerequisitos", "Procedimiento", "Próximos Pasos",
                     "Description", "Detail Page", "Link"}
        width = 60 if name in long_cols else 24
        ws.column_dimensions[get_column_letter(c)].width = width
    ws.freeze_panes = "A2"


def main() -> int:
    batches = sorted(ENRICHED_BATCHES_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime)
    if not batches:
        print(f"No batch files in {ENRICHED_BATCHES_DIR} — nothing to merge.")
        return 1
    print(f"Merging {len(batches)} batch files from {ENRICHED_BATCHES_DIR}")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    summary = []
    for sheet in ENRICHED_SHEETS:
        cols, rows = _merge_sheet(sheet, batches)
        if not cols:
            # Create empty sheet with no headers (rare; happens if every batch lacks it)
            wb.create_sheet(title=sheet)
            summary.append((sheet, 0))
            continue
        ws = wb.create_sheet(title=sheet)
        _write_sheet(ws, cols, rows)
        summary.append((sheet, len(rows)))
    ENRICHED_FINAL.parent.mkdir(parents=True, exist_ok=True)
    wb.save(ENRICHED_FINAL)
    print(f"Wrote {ENRICHED_FINAL}")
    for sheet, n in summary:
        print(f"  {sheet}: {n} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
